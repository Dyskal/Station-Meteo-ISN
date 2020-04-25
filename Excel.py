from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from os import getcwd, path
from Data import getData


def createXlsx():  # Cette fonction permet de créer le fichier excel
    titres = ['Date', 'Température', 'Pression', 'Humidité']  # On définit les titres
    pos = 0
    excel = Workbook()
    page = excel.active

    print('Fusion des cellules...')  # On fusionne les cellules de titre
    page.merge_cells('A1:B2')
    page.merge_cells('C1:D2')
    page.merge_cells('E1:F2')
    page.merge_cells('G1:H2')

    print('Création des titres...')  # On crée les titres dans les cellules fusionnées
    for i in range(1, 9, 2):
        cell = page.cell(row=1, column=i)
        cell.value = titres[pos]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        pos += 1
    page['Z1'] = 3

    excel.save(getcwd() + '\\python_weather_data.xlsx')
    excel.close()
    print('Fichier créé !')


def updateXslx():  # Cette fonction permet de mettre à jour le fichier excel
    data = getData()
    global excelwb
    page = excelwb.active
    line = page['Z1']

    page.merge_cells('A' + str(line.value) + ':' + 'B' + str(line.value))
    page.merge_cells('C' + str(line.value) + ':' + 'D' + str(line.value))
    page.merge_cells('E' + str(line.value) + ':' + 'F' + str(line.value))
    page.merge_cells('G' + str(line.value) + ':' + 'H' + str(line.value))
    page['A' + str(line.value)] = data["Time"]
    page['C' + str(line.value)] = data["Temperature"]
    page['E' + str(line.value)] = data["Pressure"]
    page['G' + str(line.value)] = data["Humidity"]
    page['Z1'] = line.value + 1

    excelwb.save(getcwd() + '\\python_weather_data.xlsx')
    print('Mise à jour effectuée !')


if not path.exists(getcwd() + '\\python_weather_data.xlsx'):  # On vérifie si le fichier excel existe
    createXlsx()  # Sinon on le crée
    excelwb = load_workbook(getcwd() + '\\python_weather_data.xlsx')
else:
    excelwb = load_workbook(getcwd() + '\\python_weather_data.xlsx')
