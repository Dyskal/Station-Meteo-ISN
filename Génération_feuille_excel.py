from openpyxl import *
from openpyxl.styles import *

titres = ['Time', 'Temperature', 'Pressure', 'Humidity']
pos = 0
excel = Workbook()
page = excel.active

print('Merging cells...')
page.merge_cells('A1:B2')
page.merge_cells('C1:D2')
page.merge_cells('E1:F2')
page.merge_cells('G1:H2')

print('Making titles...')
for i in range(1, 9, 2):
    cell = page.cell(row=1, column=i)
    cell.value = titres[pos]
    cell.alignment = Alignment(horizontal='center', vertical='center')
    pos += 1

excel.save(':\\python_weather_data.xlsx')
print('All done !')
