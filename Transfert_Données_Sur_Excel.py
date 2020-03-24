import openpyxl
from openpyxl.styles import *
from time import sleep
#from sense_emu import SenseHat
from datetime import datetime
from collections import OrderedDict

#sense = SenseHat()

data = OrderedDict()
i = 3

#récupération des données

while True:
    #if appuie_joystic = True:
    if int(input('Mettre à jour les données ? 1.Oui 2.Non\n')) == 1:
        data["Time"] = datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
        data["Temperature"] = 19
        data["Humidity"] = 50
        data["Pressure"] = 1024
        """
        data["Time"] = datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
        data["Température"] = round(sense.temp, 2)
        data["Pression"] = round(sense.pressure, 2)
        data["Humidité"] = round(sense.humidity, 2)
        """
#intégration dans la page excel
        excel = load_workbook('C:\\Weather\\python_weather_datas.xlsx')
        page = excel.active

        page.merge_cells('A'+str(i)+':'+'B'+str(i))
        page.merge_cells('C'+str(i)+':'+'D'+str(i))
        page.merge_cells('E'+str(i)+':'+'F'+str(i))
        page.merge_cells('G'+str(i)+':'+'H'+str(i))
        page['A'+str(i)] = data["Time"]
        page['C'+str(i)] = data["Temperature"]
        page['E'+str(i)] = data["Humidity"]
        page['G'+str(i)] = data["Pressure"]
        i += 1
        excel.save(':\\python_weather_datas.xlsx')
        print('Mise à jour éffectuée !')




